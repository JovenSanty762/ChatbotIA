"""
cups_referencia.py — Validación de códigos de procedimiento CUPS.

Verifica que el código de procedimiento leído de la orden médica exista en la
tabla oficial CUPS (TablaReferencia_CUPS__1.xlsx). Para no depender de openpyxl
ni parsear el Excel en cada arranque, la lista de códigos se mantiene en el
archivo de texto `cups_codigos.txt` (un código por línea, solo dígitos), que se
carga UNA vez en memoria (cache).

Regenerar `cups_codigos.txt` cuando cambie el Excel:
    python cups_referencia.py --regenerar
"""
import os
import re

_DIR = os.path.dirname(os.path.abspath(__file__))
_ARCHIVO_TXT = os.path.join(_DIR, "cups_codigos.txt")
_ARCHIVO_XLSX = os.path.join(_DIR, "TablaReferencia_CUPS__1.xlsx")

_codigos_cache = None  # set[str] con códigos normalizados (solo dígitos) o None


def _normalizar(codigo) -> str:
    """Deja solo los dígitos del código (p. ej. '86.14.02' → '861402')."""
    return re.sub(r"\D", "", str(codigo or ""))


def _cargar() -> set:
    """Carga los códigos a memoria (una sola vez). Prioriza el .txt; si no está,
    intenta el Excel con openpyxl. Devuelve un set (posiblemente vacío)."""
    global _codigos_cache
    if _codigos_cache is not None:
        return _codigos_cache

    codigos = set()
    # 1) Archivo de texto (rápido, sin dependencias)
    if os.path.exists(_ARCHIVO_TXT):
        try:
            with open(_ARCHIVO_TXT, encoding="utf-8") as f:
                for linea in f:
                    c = _normalizar(linea)
                    if c:
                        codigos.add(c)
        except Exception as e:
            print(f"⚠️ No se pudo leer cups_codigos.txt: {e}")

    # 2) Respaldo: Excel (si el .txt no existe o quedó vacío)
    if not codigos and os.path.exists(_ARCHIVO_XLSX):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_ARCHIVO_XLSX, read_only=True, data_only=True)
            ws = wb["Table"] if "Table" in wb.sheetnames else wb.worksheets[0]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                c = _normalizar(row[1] if len(row) > 1 else "")
                if c:
                    codigos.add(c)
            wb.close()
        except Exception as e:
            print(f"⚠️ No se pudo leer la tabla CUPS del Excel: {e}")

    if not codigos:
        print("⚠️ Tabla CUPS no disponible: la validación de código quedará deshabilitada.")
    _codigos_cache = codigos
    return codigos


def tabla_disponible() -> bool:
    """True si hay códigos CUPS cargados (para saber si se puede validar)."""
    return len(_cargar()) > 0


def existe(codigo) -> bool:
    """
    True si el código de procedimiento existe en la tabla CUPS. Tolera separadores
    y ceros a la izquierda faltantes (prueba también con relleno a 6 dígitos).
    IMPORTANTE: si la tabla no está disponible, devuelve True (no bloquea) para no
    detener el agendamiento por falta del archivo de referencia.
    """
    codigos = _cargar()
    if not codigos:
        return True  # tabla no disponible → no se puede validar, no bloquear
    c = _normalizar(codigo)
    if not c:
        return False
    if c in codigos:
        return True
    # Algunos CUPS son de 6 dígitos con ceros a la izquierda que el OCR puede perder.
    return c.zfill(6) in codigos


def _regenerar() -> None:
    """Reescribe cups_codigos.txt a partir del Excel (uso manual)."""
    import openpyxl
    wb = openpyxl.load_workbook(_ARCHIVO_XLSX, read_only=True, data_only=True)
    ws = wb["Table"] if "Table" in wb.sheetnames else wb.worksheets[0]
    codigos = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        c = _normalizar(row[1] if len(row) > 1 else "")
        if c:
            codigos.add(c)
    wb.close()
    with open(_ARCHIVO_TXT, "w", encoding="utf-8") as f:
        f.write("\n".join(sorted(codigos)) + "\n")
    print(f"✅ cups_codigos.txt regenerado: {len(codigos)} códigos")


if __name__ == "__main__":
    import sys
    if "--regenerar" in sys.argv:
        _regenerar()
    else:
        print(f"Códigos CUPS cargados: {len(_cargar())}")
        for prueba in ("861402", "895101", "000000", "89-51-01"):
            print(f"  existe({prueba!r}) = {existe(prueba)}")
